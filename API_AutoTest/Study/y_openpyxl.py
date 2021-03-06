import openpyxl
from openpyxl import Workbook
import time


class ExcelOpenPyxl(object):
    """
    初始化数据
    """
    def __init__(self):
        self.book = Workbook()  # 创建一个新的 xlsx 文件
        self.sheet = self.book.active  # 该函数调用工作表的索引(_active_sheet_index)，默认是0.除非修改该值，不然一直对第一张表进行操作
        # self.sheet.title = 'New title'  # 通过该属性修改工作表的名称
        # self.sheet.sheet_properties.tabColor = '1072BA'  # 提供一个RRGGBB颜色码改变标签栏的字体颜色
        # self.sheets = self.book.sheetnames()  # 获取所有的sheet页


    def write_workbook(self):

        self.sheet['A1'] = 57  # 第一种方式,写入操作--赋值
        self.sheet['A2'] = 43

        now = time.strftime("%x")
        self.sheet['A3'] = now

        self.sheet.cell(row=2, column=2).value = 2  # 第二种写入操作

        self.book.save("sample.xlsx")  # 保存操作 特别警告：这个操作将会在没有任何提示的情况下用现在写的内容，覆盖掉原文件中的所有内容

    def new_create_workbook(self, title, location=None):
        """
        创建新的sheet页
        :param title: 新创建的sheet名称
        :param location: 新创建的sheet位置
        :return:
        """
        self.title = title
        self.location = location
        self.book.create_sheet(self.title, self.location)  # 默认插在工作薄末尾
        self.book.save("sample.xlsx")

    def get_sheet_properties(self):
        """
        sheet相关属性
        :param sheet_name: sheet名称
        :return:
        """
        self.sheet.sheet_properties.tabColor = '1072BA'
        self.sheet.title = 'Yolo'
        # self.book['Yolo']  # 或者这种方式获取sheet
        # xls_gb = self.book.get_sheet_by_name('Yolo')  # 通过索引获取sheet

        self.book.save("sample.xlsx")

    def open_workbook(self, filepath):

        self.filepath = filepath
        xls_lw = openpyxl.load_workbook(self.filepath)  # 打开一个已经存在的工作薄
        return xls_lw





if __name__ == '__main__':
    excel = ExcelOpenPyxl()
    excel.write_workbook()
    # excel.new_create_workbook('User')
    print(excel.open_workbook(r'C:\Users\yolo\Desktop\Test_Case.xlsx'))


