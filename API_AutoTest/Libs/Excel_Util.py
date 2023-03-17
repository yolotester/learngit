import xlrd
import json
import pandas as pd
import openpyxl
from Libs.Log_Util import logger
from datetime import datetime
from Libs.Errors import *
from xlrd import xldate_as_tuple


class ExcelUtil(object):
    def __init__(self, filepath, sheet_name=''):
        """
        ExcelUtil构造方法
        初始化成员变量
        :param filepath:excel文件路径
        :param sheet_name: 指定sheet页名称
        """
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.xls_wb = xlrd.open_workbook(self.filepath)  # 文件名以及路径，打开xls文件
        self.sheetnames = self.xls_wb.sheet_names()  # 返回workbook中所有的sheet页的名字

    def read_excel_xlsx(self):
        """
        获取测试数据
        :return:
        """
        try:
            xls_ws = self.xls_wb.sheet_by_name(self.sheet_name)  # 通过sheet名称获取sheet
            last_row = xls_ws.nrows  # 返回该sheet中有效行数
            col_num = xls_ws.ncols  # 返回该sheet中有效列数
            logger.info("该sheet中有效行数为：%d" % last_row)
            logger.info("该sheet中有效列数为：%d" % col_num)
            if last_row > 1:
                keys = xls_ws.row_values(0)  # 返回由该行中所有单元格的数据组成的列表
                logger.info("第一行中（表头）所有单元格的数据列表：%s" % keys)
                result = []
                for col in range(1, xls_ws.nrows):
                    values = xls_ws.row_values(col)
                    logger.debug("每一行中所有单元格的数据列表：%s" % values)
                    row_dict = dict(zip(keys, values))  # zip() 函数用于将可迭代的对象作为参数，将对象中对应的元素打包成一个个元组.将表头与数据对应上打包成一个元祖
                    logger.info("将表头与数据对应，转换为字典：%s" % row_dict)
                    result.append(row_dict)
                return result

            else:
                logger.warning("【Frame】表格未填写数据，请检查！")
                return None
        except Exception as msg:
            logger.error("【Frame】读取文件异常！", msg)

    def excel_field_name_value_match(self, r_name, c_value):
        """
        匹配字段值在数据表内的第几行
        :param r_name: 字段名
        :param c_value: 字段值
        :return:
        """
        data_info = self.read_excel_xlsx()

        for row in range(len(data_info)):
            # row从0开始，在excel表中是第二行
            if data_info[row][r_name] == c_value:
                return row + 2

        logger.error('Excel中未匹配到字段数据！{}: {} = {}'.format(self.sheet_name, r_name, c_value))
        raise NotMatchExcelDataError('Excel中未匹配到字段数据！{}: {} = {}'.format(self.sheet_name, r_name, c_value))

    def get_excel_xlsx_and_type_cast(self):
        """
        获取excel表格数据并转换类型
        :return:
        """
        r_book = xlrd.open_workbook(self.filepath)
        sheet = r_book.sheet_by_name(self.sheet_name)
        rows = sheet.nrows
        cols = sheet.ncols
        all_content = []
        for i in range(1, rows):
            row_content = {}
            for j in range(cols):
                c_type = sheet.cell(i, j).ctype  # 返回的是表格的数据类型
                cell = sheet.cell_value(i, j)  # 返回的是单元格中的数据
                # logger.info("cell的数据为：%s" % cell)
                #  0. empty（空的）,1 string（text）, 2 number, 3 date, 4 boolean, 5 error， 6 blank（空白表格）
                if c_type == 2 and cell % 1 == 0:  # 如果是整型
                    cell = int(cell)
                elif c_type == 3:
                    # 转成datetime对象
                    date = datetime(*xldate_as_tuple(cell, 0))
                    cell = date.strftime("%Y/%m/%d %H:%M:%S")
                elif c_type == 4:  # 如果是布尔类型
                    cell = True if cell == 1 else False
                row_content.update({sheet.cell_value(0, j): cell})
                logger.info("行内容字典格式：%s" % row_content)
            all_content.append(row_content)
            logger.info("列表格式的所有数据：%s" % all_content)

    def write_excel(self, row_n, col_n, value):
        """
        写入测试数据
        :param row_n: 字段名
        :param col_n: 字段值
        :param value: 要写入的值
        :return:
        """
        try:
            xls_wb = openpyxl.load_workbook(self.filepath)  # openpyxl.load_workbook()来打开一个已经存在的工作簿
            xls_ws = xls_wb[self.sheet_name]  # 打开sheet页
            xls_ws.cell(row_n, col_n).value = value  # 把数据存储到某一行某一列的单元格
            xls_wb.save(self.filepath)
        except Exception as err:
            raise err

    def excel_interface_info(self, r_c_info):
        """
        获取接口模板信息
        :param r_c_info:列.行信息字典 如：{'ModuleName': '检查', 'InterfaceName': '检查API地址是否正常'} 将匹配两列的值
        :return:
        """
        data = pd.read_excel(self.filepath)
        # 匹配两个excel列的值是否存在
        if not data.loc[
            (data['ModuleName'] + data['InterfaceName'] == r_c_info['ModuleName'] + r_c_info['InterfaceName'])].empty:
            template_content = data.loc[
                (data['ModuleName'] + data['InterfaceName'] == r_c_info['ModuleName'] + r_c_info['InterfaceName'])]
            if len(template_content) > 1:
                # json.loads -- 将已编码的 JSON 字符串解码为 Python 对象
                logger.error(json.loads(template_content.to_json(orient='records')))  # 使用'records'格式化的JSON 对数据框进行编码/解码
                raise MultipleApiTempletIsFound('匹配到多个接口模板，请先确定接口模板的唯一性！')

            return json.loads(template_content.to_json(orient='records'))
        else:
            raise ApiTempletNotFound('Excel接口模板文件中未找到待匹配的接口！')

    def excel_case_info(self, sheet_name_list=[], **r_c_info):
        """
        获取接口用例/数据信息
        :param sheet_name_list: 指定sheet名称列表
        :param r_c_info: 过滤的列.行信息字典（为空则表示获取全部数据） 如：{'filter_field': {'CaseName': '检查Api地址1'}}
        :return:
        """
        result_data = []

        if isinstance(sheet_name_list, list) and sheet_name_list:
            data_dic = pd.read_excel(self.filepath, sheet_name=sheet_name_list)  # 获取指定sheet数据
        else:
            data_dic = pd.read_excel(self.filepath, sheet_name=None)  # 获取表格内所有sheet数据

        for data_ in data_dic:
            query_str = ''  # 初始化pandas查询字符串
            data = data_dic[data_]
            if r_c_info:
                for key, value in r_c_info.items():
                    if isinstance(r_c_info[key], dict):
                        for key_, value_ in r_c_info[key].items():  # 拼接pandas查询字符串
                            query_str += str(key_) + ' == "' + str(value_) + '" & '

                        query_str = query_str.rstrip(' &')
                        if not data.query(query_str).empty:
                            data = data.query(query_str)
                            result_data += json.loads(data.to_json(orient='records'))
                    else:
                        logger.error('提供的过滤运行规则字段必须为字典类型！')
                        raise TypeError('提供的过滤运行规则字段必须为字典类型！')
            else:  # 如果没有过滤字段，则查询所有数据
                result_data += json.loads(data.to_json(orient='records'))

        # 遍历数据，将部分数据字段转成字典类型
        for result in result_data:
            for key in result:
                if key in ['NewVerifParmsData', 'ExpectResult']:
                    result.update({key: json.loads(result[key]) if result[key] else {}})

        return result_data


if __name__ == '__main__':
    excel = ExcelUtil(r'C:\Users\yolo\Desktop\Test_Case.xlsx', 'User')
    # excel.get_excel_xlsx_and_type_cast()
    # excel.write_excel(2, 4, "yolo")
    data = excel.excel_interface_info({'ModuleName': '检查', 'InterfaceName': '检查API地址是否正常'})
    print(data)
    print(excel.excel_case_info(['VIP']))

    # print(excel.excel_case_info('VIP'))
    # logger.info(excel.read_excel_xlsx())
    result = excel.excel_field_name_value_match('ID', 'A001')
    print(result)
